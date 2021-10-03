#Clark P. Necciai

import spotipy as spotify
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from spotipy.oauth2 import SpotifyClientCredentials

#Identifiers necessary to utilize spotify API
cid = '7c7aa14d1c08482f9df34d7a732255e2'
secret = 'f718d8388dd64c7b8bc0f699be40752d'

client_credentials_manager = SpotifyClientCredentials(client_id=cid, client_secret=secret)

#Create microsoft excel using openpyxl
workbook = Workbook()
workbook.save(filename="KeysAndSongs.xlsx")
ws = workbook.active
ws.title = "Songs And Attributes"

#Access Spotify Credentials/API
sp = spotify.Spotify(client_credentials_manager = client_credentials_manager)

#Add first row of titular characteristics
ws.append(["Artist/Composer", "Title", "Length", "Key", "Tempo","Time Signature", "Release Date", "Energy", "Loudness", "Valence", "Danceability", "Acousticness"])

#Convert to readable time for table
def convertMillis(millis):
    seconds=(millis/1000)%60
    seconds = int(seconds)
    if (seconds < 10):
        seconds = str("0") + str(seconds)
    minutes=(millis/(1000*60))%60
    minutes = int(minutes)
    time = str(minutes) + ":"+ str(seconds)
    return time

#Spotify API utilizes numerical system instead of keys
#Built a quick table for conversions during read in from API
def detKey(keyValue):
    switcher = {
    0: "C",
    1: "C#",
    2: "D",
    3: "D#",
    4: "E",
    5: "F",
    6: "F#",
    7: "G",
    8: "G#",
    9: "A",
    10: "A#",
    11: "B",
    12: "B#",
    }
    return str(switcher.get(keyValue, "Not Known"))

#Number of songs decision by user
songCount = int(input("How many songs: "))

#Access list features from 0-songcount and return Attributes of songs
#of any playlist --> playlist ID can be changed to change playlist desired
# Ability to utilize own playlist possible
for i in range(0, songCount):
    playlist_items_list = sp.playlist_items(playlist_id = '2LOxEzC4KmoWJ9NhW0kz5M', fields = 'items', limit = 1, offset = i)

    features = sp.audio_features(playlist_items_list['items'][0]['track']['id'])
    energy = features[0]['energy']
    danceability = features[0]['danceability']
    valence = features[0]['valence']
    acousticness = features[0]['acousticness']
    instrumentalness = features[0]['instrumentalness']
    loudness = features[0]['loudness']
    tempo = int(features[0]['tempo'])
    time_signature = int(features[0]['time_signature'])
    key = detKey(features[0]['key'])

    PlayListItemsFull = playlist_items_list['items'][0]['track']
    release_date = PlayListItemsFull['album']['release_date']
    bandname = PlayListItemsFull['artists'][0]['name']
    song_title = PlayListItemsFull['name']
    time = PlayListItemsFull['duration_ms']
    time = convertMillis(time)

    #Format cells with appropriate information on every iteration
    ws.cell(row = i + 2,column = 1,value = bandname)
    ws.cell(row = i + 2,column = 2,value = song_title)
    ws.cell(row = i + 2,column = 3,value = time)
    ws.cell(row = i + 2,column = 4,value = key)
    ws.cell(row = i + 2,column = 5,value = tempo)
    ws.cell(row = i + 2,column = 6,value = time_signature)
    ws.cell(row = i + 2,column = 7,value = release_date)
    ws.cell(row = i + 2,column = 8,value = energy)
    ws.cell(row = i + 2,column = 9,value = loudness)
    ws.cell(row = i + 2,column = 10,value = valence)
    ws.cell(row = i + 2,column = 11,value = danceability)
    ws.cell(row = i + 2,column = 12,value = acousticness)
    #End range

#Save and return notebook in same file
workbook.save(filename="KeysAndSongs.xlsx")
